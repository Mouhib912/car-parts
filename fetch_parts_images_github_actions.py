"""
fetch_parts_images_github_actions.py
Downloads images from eBay with intelligent threading and rate limiting.
Adapted for GitHub Actions with batch processing (start_index, end_index).
"""

import re
import time
import requests
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from queue import Queue, Empty
from threading import Lock
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from PIL import Image as PILImage
from urllib.parse import quote
import random
import argparse
import os

# ── CONFIG ───────────────────────────────────────────────────────────────────
INPUT_FILE         = "merged_corrected_final_FIXED.xlsx"
OUTPUT_FILE_PREFIX = "results_batch"
MAX_IMG_SIZE       = (150, 150)
RETRY_ATTEMPTS     = 3                 # Retry failed requests

# Detect if running inside GitHub Actions
IS_GITHUB_ACTIONS = os.environ.get("GITHUB_ACTIONS", "false").lower() == "true"

if IS_GITHUB_ACTIONS:
    # GitHub Actions: fresh IP per runner, but be polite to avoid batch-level bans
    MAX_WORKERS = 1                    # Sequential within each batch runner
    MIN_DELAY   = 10.0                  # Minimum wait between searches (seconds)
    MAX_DELAY   = 15.0                  # Maximum wait between searches (seconds)
    print("[ENV] Running in GitHub Actions — using conservative delays (3–7s)")
else:
    # Local: shared IP, keep it slow to avoid bans
    MAX_WORKERS = 1                    # Single worker to be extra gentle
    MIN_DELAY   = 10.0                  # Minimum wait between searches (seconds)
    MAX_DELAY   = 15.0                  # Maximum wait between searches (seconds)
    print("[ENV] Running locally — using safe delays (3–6s)")
# ─────────────────────────────────────────────────────────────────────────────

# Enhanced headers to avoid bot detection
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "DNT": "1",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Cache-Control": "max-age=0",
}

# Regex for eBay images
EBAY_IMG_RE = re.compile(r"https://i\.ebayimg\.com/images/g/[^\"'\s<>]+\.(?:jpg|webp|png)")

# Thread-safe counter and lock
last_request_time = 0
time_lock = Lock()


# ── LOAD DATA ─────────────────────────────────────────────────────────────────
def load_items(path, start_index=None, end_index=None):
    if not os.path.exists(path):
        print(f"Error: Input file {path} not found.")
        return pd.DataFrame()
        
    raw = pd.read_excel(path, header=None)
    df = raw[1:].reset_index(drop=True)
    
    # Col 0=code, 1=French name (Libellé), 2=English name, 4=brand, 7=model, 8=category
    df2 = df.iloc[:, [0, 1, 2, 4, 7, 8]].copy()
    df2.columns = ["code", "name_fr", "name_en", "brand", "model", "category"]
    
    df2["name_fr"] = df2["name_fr"].astype(str).str.strip()
    df2["name_en"] = df2["name_en"].astype(str).str.strip()
    
    # Use English name when available, fall back to French Libellé
    def pick_name(row):
        en = row["name_en"]
        fr = row["name_fr"]
        if en and en.lower() != "nan":
            return en
        if fr and fr.lower() != "nan":
            return fr
        return ""
    
    df2["name"] = df2.apply(pick_name, axis=1)
    df2["name_source"] = df2.apply(
        lambda r: "EN" if r["name_en"] and r["name_en"].lower() != "nan" else "FR", axis=1
    )
    
    # Drop rows with no name at all
    df2 = df2[df2["name"] != ""].reset_index(drop=True)
    df2 = df2[["code", "name", "name_source", "brand", "model", "category"]]
    
    en_count = (df2["name_source"] == "EN").sum()
    fr_count = (df2["name_source"] == "FR").sum()
    print(f"  -> {len(df2)} total products ({en_count} English names, {fr_count} French fallbacks)")
    
    # Apply batch slicing if provided
    if start_index is not None and end_index is not None:
        df2 = df2.iloc[start_index:end_index].reset_index(drop=True)
    
    return df2


# ── RATE LIMITING ─────────────────────────────────────────────────────────────
def wait_before_request():
    """Implement intelligent rate limiting with random jitter."""
    global last_request_time
    
    with time_lock:
        elapsed = time.time() - last_request_time
        delay = random.uniform(MIN_DELAY, MAX_DELAY)
        
        if elapsed < delay:
            sleep_time = delay - elapsed
            time.sleep(sleep_time)
        
        last_request_time = time.time()


# ── EBAY IMAGE SEARCH ─────────────────────────────────────────────────────────
def search_ebay_image(name, attempt=1):
    """Return (img_url, img_data) for the first eBay image matching `name`."""
    
    # Rate limiting before request
    wait_before_request()
    
    url = f"https://www.ebay.com/sch/i.html?_nkw={quote(name)}&_sacat=6030"
    
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        
        # Check for rate limiting or blocking
        if resp.status_code == 429:
            print(f"    [429] Rate limited. Waiting 30 seconds...")
            time.sleep(30)
            if attempt < RETRY_ATTEMPTS:
                return search_ebay_image(name, attempt + 1)
            return "", None
        
        if resp.status_code == 403:
            print(f"    [403] Forbidden. eBay blocked this request.")
            return "", None
        
        if resp.status_code != 200:
            print(f"    [HTTP {resp.status_code}] Unexpected status")
            if attempt < RETRY_ATTEMPTS:
                time.sleep(2)
                return search_ebay_image(name, attempt + 1)
            return "", None
        
        # Check for CAPTCHA or robot check
        if "captcha" in resp.text.lower() or "robot" in resp.text.lower():
            print(f"    [BLOCKED] CAPTCHA/Robot check detected")
            return "", None
        
        # Extract image URL
        m = EBAY_IMG_RE.search(resp.text)
        if m:
            # Prefer a larger thumbnail
            img_url = re.sub(r"s-l\d+", "s-l500", m.group(0))
            data = download_image(img_url)
            return img_url, data
        
    except requests.exceptions.Timeout:
        print(f"    [TIMEOUT] Request timed out")
        if attempt < RETRY_ATTEMPTS:
            time.sleep(2)
            return search_ebay_image(name, attempt + 1)
    except requests.exceptions.ConnectionError as e:
        print(f"    [CONNECTION] {e}")
        if attempt < RETRY_ATTEMPTS:
            time.sleep(2)
            return search_ebay_image(name, attempt + 1)
    except Exception as e:
        print(f"    [ERROR] {type(e).__name__}: {e}")
    
    return "", None


# ── IMAGE DOWNLOAD ────────────────────────────────────────────────────────────
def download_image(url, attempt=1):
    """Download and process image with retry logic."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            if attempt < RETRY_ATTEMPTS:
                time.sleep(1)
                return download_image(url, attempt + 1)
            return None
        
        img = PILImage.open(BytesIO(resp.content)).convert("RGB")
        img.thumbnail(MAX_IMG_SIZE, PILImage.LANCZOS)
        buf = BytesIO()
        img.save(buf, format="JPEG", quality=85)
        buf.seek(0)
        return buf
        
    except Exception as e:
        print(f"    [IMG_ERROR] {type(e).__name__}: {e}")
        if attempt < RETRY_ATTEMPTS:
            time.sleep(1)
            return download_image(url, attempt + 1)
        return None


# ── WORKER FUNCTION ──────────────────────────────────────────────────────────
def fetch_image_for_part(row_dict):
    """Worker function for thread pool."""
    name = row_dict["name"]
    img_url, img_data = search_ebay_image(name)
    status = "Found" if img_data else "Missing"
    
    return {
        **row_dict,
        "img_url": img_url,
        "img_data": img_data,
        "status": status,
    }


# ── EXCEL BUILDER ─────────────────────────────────────────────────────────────
def build_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Car Parts"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    alt_fill = PatternFill("solid", start_color="D9E1F2")

    headers    = ["#", "Code", "Part Name", "Brand", "Model", "Category", "Image", "Status"]
    col_widths = [4,    15,     35,           15,      20,      22,          24,      10]

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 30

    # Data rows
    for i, r in enumerate(rows, start=1):
        row_num = i + 1
        ws.row_dimensions[row_num].height = 115
        fill = alt_fill if i % 2 == 0 else None

        values = [i, r["code"], r["name"], r["brand"], r["model"], r["category"], "", r["status"]]
        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = left_wrap if col in (3, 4, 5, 6) else center
            cell.border    = thin
            if fill:
                cell.fill  = fill

        if r["img_data"]:
            xl_img        = XLImage(r["img_data"])
            xl_img.width  = 130
            xl_img.height = 100
            ws.add_image(xl_img, f"G{row_num}")
        else:
            ws.cell(row=row_num, column=7, value="No image")

    wb.save(output_path)
    return output_path


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="eBay Scraper for GitHub Actions")
    parser.add_argument("--start-index", type=int, help="Start index of rows to process")
    parser.add_argument("--end-index", type=int, help="End index of rows to process")
    args = parser.parse_args()

    global last_request_time
    
    t0 = time.time()
    print(f"Loading: {INPUT_FILE}")
    df = load_items(INPUT_FILE, args.start_index, args.end_index)
    total = len(df)
    
    if total == 0:
        print("No items to process. Exiting.")
        return

    print(f"  -> {total} parts to process in this batch")
    print(f"  -> Using {MAX_WORKERS} workers with {MIN_DELAY}-{MAX_DELAY}s delays\n")

    rows = []
    found = 0
    
    # Initialize last request time
    last_request_time = time.time()

    # Use ThreadPoolExecutor with controlled concurrency
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_image_for_part, row.to_dict()): idx 
                   for idx, (_, row) in enumerate(df.iterrows())}
        
        for future in as_completed(futures):
            idx = futures[future]
            try:
                result = future.result()
                rows.append(result)
                
                if result["img_data"]:
                    found += 1
                
                num = idx + 1
                status_icon = "✓" if result["img_data"] else "✗"
                print(f"[{num}/{total}] {status_icon} {result['name'][:60]}")
                
            except Exception as e:
                print(f"[{idx+1}/{total}] ✗ Error: {e}")
                rows.append({
                    **df.iloc[idx].to_dict(),
                    "img_url": "",
                    "img_data": None,
                    "status": "Error",
                })

    # Sort results by original dataframe order
    rows_sorted = sorted(rows, key=lambda x: list(df["code"]).index(x["code"]) 
                        if x["code"] in list(df["code"]) else float('inf'))
    
    # Final output
    output_file = f"{OUTPUT_FILE_PREFIX}_{args.start_index}_{args.end_index}.xlsx"
    print(f"\nBuilding batch Excel -> {output_file}")
    build_excel(rows_sorted, output_file)
    elapsed = time.time() - t0
    pct = found / total * 100 if total else 0
    print(f"\nDone! {found}/{total} images embedded ({pct:.0f}%) in {elapsed/60:.1f} minutes.")
    print(f"Output saved to: {output_file}")


if __name__ == "__main__":
    main()
